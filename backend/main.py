from __future__ import annotations

import base64
import hashlib
import hmac
import json
import os
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Literal

from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = Path(__file__).resolve().parent / "data"
DB_PATH = DATA_DIR / "app.db"
LAYERS_JSON_PATH = PROJECT_ROOT / "public" / "geodata" / "layers.json"
MOCK_SMS_CODE = "123456"

AuthStatus = Literal["unregistered", "pending", "approved", "disabled", "rejected", "supplement"]
UserType = Literal["系统管理员", "居民", "政府管理者", "市场主体"]
MapTopic = Literal["problem", "demand", "parcel"]


class PasswordLoginPayload(BaseModel):
  username: str
  password: str


class SmsPayload(BaseModel):
  phone: str


class SmsLoginPayload(BaseModel):
  phone: str
  code: str


class RegistrationPayload(BaseModel):
  phone: str
  name: str
  userType: UserType
  contact: str
  email: str | None = None
  password: str | None = None
  street: str | None = None
  parcel: str | None = None
  address: str | None = None


class WhitelistUserPayload(BaseModel):
  phone: str
  name: str
  userType: UserType
  street: str | None = None
  parcel: str | None = None
  address: str | None = None
  accessScope: str | None = None
  allowFeedback: bool = True
  allowManage: bool = False
  status: AuthStatus = "unregistered"
  contact: str | None = None
  email: str | None = None


class UserStatusPayload(BaseModel):
  status: AuthStatus


class LayerPatchPayload(BaseModel):
  name: str | None = None
  categoryName: str | None = None
  enabled: bool | None = None
  sortOrder: int | None = None
  note: str | None = None


app = FastAPI(title="南宁城市体检信息平台本地后端")

app.add_middleware(
  CORSMiddleware,
  allow_origins=[
    "http://127.0.0.1:5173",
    "http://localhost:5173",
    "http://127.0.0.1:5174",
    "http://localhost:5174",
    "http://127.0.0.1:4173",
    "http://localhost:4173",
  ],
  allow_origin_regex=r"^http://(127\.0\.0\.1|localhost):\d+$",
  allow_credentials=True,
  allow_methods=["*"],
  allow_headers=["*"],
)


def now_iso() -> str:
  return datetime.now(timezone.utc).isoformat()


def get_connection() -> sqlite3.Connection:
  DATA_DIR.mkdir(parents=True, exist_ok=True)
  connection = sqlite3.connect(DB_PATH)
  connection.row_factory = sqlite3.Row
  return connection


def hash_password(password: str) -> str:
  iterations = 200_000
  salt = os.urandom(16)
  digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
  return "pbkdf2_sha256${}${}${}".format(
    iterations,
    base64.b64encode(salt).decode("ascii"),
    base64.b64encode(digest).decode("ascii"),
  )


def verify_password(password: str, stored_hash: str | None) -> bool:
  if not stored_hash:
    return False
  try:
    algorithm, iterations_text, salt_text, digest_text = stored_hash.split("$", 3)
    if algorithm != "pbkdf2_sha256":
      return False
    expected = hashlib.pbkdf2_hmac(
      "sha256",
      password.encode("utf-8"),
      base64.b64decode(salt_text),
      int(iterations_text),
    )
    return hmac.compare_digest(expected, base64.b64decode(digest_text))
  except (ValueError, TypeError):
    return False


def bool_from_db(value: Any) -> bool:
  return bool(int(value or 0))


def normalized_text(value: Any) -> str:
  if value is None:
    return ""
  return str(value).strip()


def normalized_phone(value: Any) -> str:
  if value is None:
    return ""
  if isinstance(value, float) and value.is_integer():
    return str(int(value))
  return str(value).strip()


def parse_bool(value: Any, default: bool = False) -> bool:
  if value is None or value == "":
    return default
  if isinstance(value, bool):
    return value
  if isinstance(value, (int, float)):
    return value != 0
  return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "启用", "允许", "可"}


def normalize_user_type(value: Any) -> str:
  text = normalized_text(value)
  if text in {"系统管理员", "居民", "政府管理者", "市场主体"}:
    return text
  return "居民"


def normalize_status(value: Any) -> str:
  text = normalized_text(value)
  mapping = {
    "未注册": "unregistered",
    "待审核": "pending",
    "已通过": "approved",
    "已禁用": "disabled",
    "审核驳回": "rejected",
    "需补充材料": "supplement",
  }
  if text in {"unregistered", "pending", "approved", "disabled", "rejected", "supplement"}:
    return text
  return mapping.get(text, "unregistered")


def user_to_public(row: sqlite3.Row) -> dict[str, Any]:
  return {
    "phone": row["phone"],
    "name": row["name"],
    "userType": row["user_type"],
    "accessScope": row["access_scope"],
    "allowFeedback": bool_from_db(row["allow_feedback"]),
    "allowManage": bool_from_db(row["allow_manage"]),
  }


def user_to_admin(row: sqlite3.Row) -> dict[str, Any]:
  data = user_to_public(row)
  data.update(
    {
      "id": row["id"],
      "username": row["username"],
      "street": row["street"],
      "parcel": row["parcel"],
      "address": row["address"],
      "status": row["status"],
      "contact": row["contact"],
      "email": row["email"],
      "createdAt": row["created_at"],
      "updatedAt": row["updated_at"],
    }
  )
  return data


def layer_to_public(row: sqlite3.Row) -> dict[str, Any]:
  return {
    "id": row["id"],
    "name": row["name"],
    "category": row["category"],
    "categoryName": row["category_name"],
    "geometryType": row["geometry_type"],
    "featureCount": row["feature_count"],
    "crs": row["crs"],
    "file": row["file"],
    "objectType": row["object_type"],
    "diagnosisDimension": row["diagnosis_dimension"],
    "enabled": bool_from_db(row["enabled"]),
    "sortOrder": row["sort_order"],
    "note": row["note"],
  }


def upsert_whitelist_user(connection: sqlite3.Connection, payload: dict[str, Any]) -> tuple[sqlite3.Row, bool]:
  phone = normalized_phone(payload.get("phone"))
  if not phone:
    raise ValueError("手机号不能为空")
  name = normalized_text(payload.get("name"))
  if not name:
    raise ValueError("姓名不能为空")

  existing = connection.execute("SELECT * FROM users WHERE phone = ?", (phone,)).fetchone()
  timestamp = now_iso()
  values = {
    "phone": phone,
    "name": name,
    "user_type": normalize_user_type(payload.get("userType") or payload.get("user_type")),
    "street": normalized_text(payload.get("street")),
    "parcel": normalized_text(payload.get("parcel")),
    "address": normalized_text(payload.get("address")),
    "access_scope": normalized_text(payload.get("accessScope") or payload.get("access_scope")) or "注册审核通过后开通",
    "allow_feedback": int(parse_bool(payload.get("allowFeedback") if "allowFeedback" in payload else payload.get("allow_feedback"), True)),
    "allow_manage": int(parse_bool(payload.get("allowManage") if "allowManage" in payload else payload.get("allow_manage"), False)),
    "status": normalize_status(payload.get("status")),
    "contact": normalized_text(payload.get("contact")) or phone,
    "email": normalized_text(payload.get("email")),
  }

  if existing:
    connection.execute(
      """
      UPDATE users
      SET name = ?, user_type = ?, street = ?, parcel = ?, address = ?, access_scope = ?,
          allow_feedback = ?, allow_manage = ?, status = ?, contact = ?, email = ?, updated_at = ?
      WHERE phone = ?
      """,
      (
        values["name"],
        values["user_type"],
        values["street"],
        values["parcel"],
        values["address"],
        values["access_scope"],
        values["allow_feedback"],
        values["allow_manage"],
        values["status"],
        values["contact"],
        values["email"],
        timestamp,
        phone,
      ),
    )
    updated = connection.execute("SELECT * FROM users WHERE phone = ?", (phone,)).fetchone()
    return updated, False

  connection.execute(
    """
    INSERT INTO users (
      phone, username, password_hash, name, user_type, street, parcel, address,
      access_scope, allow_feedback, allow_manage, status, contact, email, created_at, updated_at
    ) VALUES (?, NULL, NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """,
    (
      phone,
      values["name"],
      values["user_type"],
      values["street"],
      values["parcel"],
      values["address"],
      values["access_scope"],
      values["allow_feedback"],
      values["allow_manage"],
      values["status"],
      values["contact"],
      values["email"],
      timestamp,
      timestamp,
    ),
  )
  inserted = connection.execute("SELECT * FROM users WHERE phone = ?", (phone,)).fetchone()
  return inserted, True


def create_schema(connection: sqlite3.Connection) -> None:
  connection.executescript(
    """
    CREATE TABLE IF NOT EXISTS users (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      phone TEXT NOT NULL UNIQUE,
      username TEXT UNIQUE,
      password_hash TEXT,
      name TEXT NOT NULL,
      user_type TEXT NOT NULL,
      street TEXT,
      parcel TEXT,
      address TEXT,
      access_scope TEXT NOT NULL,
      allow_feedback INTEGER NOT NULL DEFAULT 0,
      allow_manage INTEGER NOT NULL DEFAULT 0,
      status TEXT NOT NULL,
      contact TEXT,
      email TEXT,
      created_at TEXT NOT NULL,
      updated_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS layers (
      id TEXT PRIMARY KEY,
      name TEXT NOT NULL,
      category TEXT NOT NULL,
      category_name TEXT NOT NULL,
      geometry_type TEXT NOT NULL,
      feature_count INTEGER NOT NULL DEFAULT 0,
      crs TEXT NOT NULL,
      file TEXT NOT NULL,
      object_type TEXT,
      diagnosis_dimension TEXT,
      enabled INTEGER NOT NULL DEFAULT 1,
      sort_order INTEGER NOT NULL DEFAULT 0,
      note TEXT,
      created_at TEXT NOT NULL,
      updated_at TEXT NOT NULL
    );
    """
  )
  connection.commit()


def seed_users(connection: sqlite3.Connection) -> None:
  count = connection.execute("SELECT COUNT(*) FROM users").fetchone()[0]
  if count:
    return

  users = [
    {
      "phone": "0771000000",
      "username": "admin",
      "password": "admin123",
      "name": "系统管理员",
      "user_type": "系统管理员",
      "street": "",
      "parcel": "",
      "address": "",
      "access_scope": "全市空间范围、指标体系、白名单与后台管理",
      "allow_feedback": 1,
      "allow_manage": 1,
      "status": "approved",
      "contact": "0771000000",
      "email": "",
    },
    {
      "phone": "13800138000",
      "username": None,
      "password": None,
      "name": "张晓明",
      "user_type": "政府管理者",
      "street": "福建园街道",
      "parcel": "JNY-02-08",
      "address": "江南中街片区",
      "access_scope": "江南区城市体检综合分析、问题一张图、需求一张图",
      "allow_feedback": 1,
      "allow_manage": 1,
      "status": "approved",
      "contact": "13800138000",
      "email": "",
    },
    {
      "phone": "13800138001",
      "username": None,
      "password": None,
      "name": "李雨",
      "user_type": "居民",
      "street": "亭洪街道",
      "parcel": "TH-01-03",
      "address": "亭洪路社区 3 栋",
      "access_scope": "本人住址及周边社区体检结果",
      "allow_feedback": 1,
      "allow_manage": 0,
      "status": "pending",
      "contact": "13800138001",
      "email": "",
    },
    {
      "phone": "13800138002",
      "username": None,
      "password": None,
      "name": "周伟",
      "user_type": "市场主体",
      "street": "白沙街道",
      "parcel": "BS-04-11",
      "address": "白沙商业街 12 号",
      "access_scope": "绑定地块与反馈统计",
      "allow_feedback": 1,
      "allow_manage": 0,
      "status": "disabled",
      "contact": "13800138002",
      "email": "",
    },
    {
      "phone": "13800138003",
      "username": None,
      "password": None,
      "name": "黄静",
      "user_type": "居民",
      "street": "江南街道",
      "parcel": "JN-07-02",
      "address": "淡村社区",
      "access_scope": "注册审核通过后开通",
      "allow_feedback": 1,
      "allow_manage": 0,
      "status": "unregistered",
      "contact": "13800138003",
      "email": "",
    },
    {
      "phone": "13800138004",
      "username": None,
      "password": None,
      "name": "陈启",
      "user_type": "政府管理者",
      "street": "沙井街道",
      "parcel": "SJ-06-05",
      "address": "沙井片区",
      "access_scope": "需补充材料后开通",
      "allow_feedback": 0,
      "allow_manage": 1,
      "status": "supplement",
      "contact": "13800138004",
      "email": "",
    },
  ]

  timestamp = now_iso()
  for user in users:
    connection.execute(
      """
      INSERT INTO users (
        phone, username, password_hash, name, user_type, street, parcel, address,
        access_scope, allow_feedback, allow_manage, status, contact, email, created_at, updated_at
      ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
      """,
      (
        user["phone"],
        user["username"],
        hash_password(user["password"]) if user["password"] else None,
        user["name"],
        user["user_type"],
        user["street"],
        user["parcel"],
        user["address"],
        user["access_scope"],
        user["allow_feedback"],
        user["allow_manage"],
        user["status"],
        user["contact"],
        user["email"],
        timestamp,
        timestamp,
      ),
    )
  connection.commit()


def ensure_demo_access_users(connection: sqlite3.Connection) -> None:
  demo_users = [
    {
      "phone": "13800138005",
      "name": "林小雅",
      "user_type": "居民",
      "street": "亭洪街道",
      "parcel": "TH-02-06",
      "address": "亭洪路社区 6 栋",
      "access_scope": "本人住址及周边社区体检结果",
      "allow_feedback": 1,
      "allow_manage": 0,
      "status": "approved",
      "contact": "13800138005",
      "email": "",
    },
    {
      "phone": "13800138006",
      "name": "南宁乐创园区",
      "user_type": "市场主体",
      "street": "白沙街道",
      "parcel": "BS-05-09",
      "address": "白沙商业街 20 号",
      "access_scope": "绑定地块与综合分析结果",
      "allow_feedback": 1,
      "allow_manage": 0,
      "status": "approved",
      "contact": "13800138006",
      "email": "",
    },
  ]

  timestamp = now_iso()
  for user in demo_users:
    existing = connection.execute("SELECT id FROM users WHERE phone = ?", (user["phone"],)).fetchone()
    if existing:
      continue
    connection.execute(
      """
      INSERT INTO users (
        phone, username, password_hash, name, user_type, street, parcel, address,
        access_scope, allow_feedback, allow_manage, status, contact, email, created_at, updated_at
      ) VALUES (?, NULL, NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
      """,
      (
        user["phone"],
        user["name"],
        user["user_type"],
        user["street"],
        user["parcel"],
        user["address"],
        user["access_scope"],
        user["allow_feedback"],
        user["allow_manage"],
        user["status"],
        user["contact"],
        user["email"],
        timestamp,
        timestamp,
      ),
    )
  connection.commit()


def seed_layers(connection: sqlite3.Connection) -> None:
  count = connection.execute("SELECT COUNT(*) FROM layers").fetchone()[0]
  if count:
    return
  if not LAYERS_JSON_PATH.exists():
    return

  layers = json.loads(LAYERS_JSON_PATH.read_text(encoding="utf-8"))
  timestamp = now_iso()
  for index, layer in enumerate(layers):
    connection.execute(
      """
      INSERT INTO layers (
        id, name, category, category_name, geometry_type, feature_count, crs, file,
        object_type, diagnosis_dimension, enabled, sort_order, note, created_at, updated_at
      ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
      """,
      (
        layer["id"],
        layer["name"],
        layer["category"],
        layer["categoryName"],
        layer["geometryType"],
        int(layer.get("featureCount", 0)),
        layer.get("crs", "EPSG:4326"),
        layer["file"],
        layer.get("objectType"),
        layer.get("diagnosisDimension"),
        1,
        index,
        "",
        timestamp,
        timestamp,
      ),
    )
  connection.commit()


def initialize_database() -> None:
  with get_connection() as connection:
    create_schema(connection)
    seed_users(connection)
    ensure_demo_access_users(connection)
    seed_layers(connection)


@app.on_event("startup")
def on_startup() -> None:
  initialize_database()


@app.get("/api/health")
def health() -> dict[str, Any]:
  return {"ok": True, "message": "本地后端运行正常。"}


@app.get("/api/auth/whitelist/{phone}")
def get_whitelist_user(phone: str) -> dict[str, Any]:
  with get_connection() as connection:
    row = connection.execute("SELECT * FROM users WHERE phone = ?", (phone.strip(),)).fetchone()
    if not row:
      return {"ok": False, "message": "该手机号不在白名单中，请联系管理员录入。"}
    return {"ok": True, "user": user_to_admin(row)}


@app.post("/api/auth/sms/send")
def send_sms_code(payload: SmsPayload) -> dict[str, Any]:
  with get_connection() as connection:
    row = connection.execute("SELECT * FROM users WHERE phone = ?", (payload.phone.strip(),)).fetchone()
    if not row:
      return {"ok": False, "message": "该手机号不在白名单中，请联系管理员录入。"}
    if row["status"] == "disabled":
      return {"ok": False, "message": "该账号已禁用，请联系平台管理员。"}
    return {"ok": True, "message": f"验证码已发送，本地演示验证码为 {MOCK_SMS_CODE}。"}


@app.post("/api/auth/login/sms")
def login_with_sms(payload: SmsLoginPayload) -> dict[str, Any]:
  with get_connection() as connection:
    row = connection.execute("SELECT * FROM users WHERE phone = ?", (payload.phone.strip(),)).fetchone()
    if not row:
      return {"ok": False, "message": "该手机号不在白名单中，请联系管理员录入。"}
    if payload.code != MOCK_SMS_CODE:
      return {"ok": False, "message": "验证码不正确，请输入本地演示验证码 123456。"}
    if row["status"] == "pending":
      return {"ok": False, "message": "注册信息正在审核中，审核通过后将开通账号。"}
    if row["status"] == "disabled":
      return {"ok": False, "message": "该账号已禁用，请联系平台管理员。"}
    if row["status"] == "unregistered":
      return {"ok": False, "message": "该手机号尚未完成注册，请先提交注册信息。"}
    if row["status"] == "supplement":
      return {"ok": False, "message": "注册信息需补充认证材料，请在注册页完善后提交。"}
    if row["status"] == "rejected":
      return {"ok": False, "message": "注册申请已被驳回，请联系管理员确认原因。"}
    return {"ok": True, "user": user_to_public(row), "message": "登录成功，正在进入平台。"}


@app.post("/api/auth/login/password")
def login_with_password(payload: PasswordLoginPayload) -> dict[str, Any]:
  identity = payload.username.strip()
  with get_connection() as connection:
    row = connection.execute(
      "SELECT * FROM users WHERE username = ? OR phone = ?",
      (identity, identity),
    ).fetchone()
    if not row or not verify_password(payload.password, row["password_hash"]):
      return {"ok": False, "message": "账号或密码不正确。管理员账号：admin / admin123。"}
    if row["status"] != "approved":
      return {"ok": False, "message": "账号未启用，请等待管理员审核或联系管理员处理。"}
    return {"ok": True, "user": user_to_public(row), "message": "管理员登录成功。"}


@app.post("/api/registrations")
def submit_registration(payload: RegistrationPayload) -> dict[str, Any]:
  phone = payload.phone.strip()
  with get_connection() as connection:
    row = connection.execute("SELECT * FROM users WHERE phone = ?", (phone,)).fetchone()
    if not row:
      return {"ok": False, "message": "该手机号不在白名单中，暂不能提交注册。"}
    if row["status"] == "disabled":
      return {"ok": False, "message": "该账号已禁用，无法提交注册。"}
    if row["status"] == "approved":
      return {"ok": False, "message": "该手机号已开通账号，可直接登录。"}

    connection.execute(
      """
      UPDATE users
      SET name = ?, user_type = ?, street = ?, parcel = ?, address = ?, contact = ?,
          email = ?, password_hash = COALESCE(?, password_hash), status = 'pending', updated_at = ?
      WHERE phone = ?
      """,
      (
        payload.name,
        payload.userType,
        payload.street or "",
        payload.parcel or "",
        payload.address or "",
        payload.contact,
        payload.email or "",
        hash_password(payload.password) if payload.password else None,
        now_iso(),
        phone,
      ),
    )
    connection.commit()
    return {
      "ok": True,
      "message": "注册信息已提交，当前状态为待审核。审核通过后会按白名单权限开通账号。",
      "status": "pending",
    }


@app.get("/api/admin/users")
def list_admin_users() -> list[dict[str, Any]]:
  with get_connection() as connection:
    rows = connection.execute("SELECT * FROM users ORDER BY id ASC").fetchall()
    return [user_to_admin(row) for row in rows]


@app.post("/api/admin/users")
def create_whitelist_user(payload: WhitelistUserPayload) -> dict[str, Any]:
  try:
    with get_connection() as connection:
      row, created = upsert_whitelist_user(connection, payload.model_dump())
      connection.commit()
      return {
        "ok": True,
        "message": "白名单用户已新增。" if created else "该手机号已存在，白名单信息已更新。",
        "user": user_to_admin(row),
        "created": created,
      }
  except ValueError as error:
    raise HTTPException(status_code=400, detail=str(error)) from error


@app.post("/api/admin/users/import")
async def import_whitelist_users(file: UploadFile = File(...)) -> dict[str, Any]:
  filename = file.filename or ""
  if not filename.lower().endswith(".xlsx"):
    raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的白名单 Excel 文件。")

  content = await file.read()
  workbook_path = DATA_DIR / "_whitelist_import.xlsx"
  DATA_DIR.mkdir(parents=True, exist_ok=True)
  workbook_path.write_bytes(content)

  field_aliases = {
    "phone": {"手机号", "手机", "电话", "联系电话", "phone"},
    "name": {"姓名", "名称", "用户姓名", "name"},
    "userType": {"用户类型", "人员类型", "角色", "userType", "user_type"},
    "street": {"街道", "街区", "所属街道", "关联街区", "street"},
    "parcel": {"地块", "关联地块", "parcel"},
    "address": {"地址", "住址", "关联住址", "关联住址/对象", "对象", "address"},
    "accessScope": {"权限范围", "访问范围", "权限", "accessScope", "access_scope"},
    "allowFeedback": {"允许反馈", "反馈权限", "allowFeedback", "allow_feedback"},
    "allowManage": {"允许管理", "管理权限", "后台权限", "allowManage", "allow_manage"},
    "status": {"状态", "审核状态", "status"},
    "contact": {"联系方式", "联系人电话", "contact"},
    "email": {"邮箱", "电子邮箱", "email"},
  }

  created_count = 0
  updated_count = 0
  skipped_rows: list[dict[str, Any]] = []

  try:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook.active
    headers = [normalized_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    column_map: dict[int, str] = {}
    for index, header in enumerate(headers):
      for field, aliases in field_aliases.items():
        if header in aliases:
          column_map[index] = field
          break

    if "phone" not in column_map.values() or "name" not in column_map.values():
      raise HTTPException(status_code=400, detail="Excel 第一行必须包含“手机号”和“姓名”列。")

    with get_connection() as connection:
      for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        payload = {field: row[index] for index, field in column_map.items() if index < len(row)}
        if not any(normalized_text(value) for value in payload.values()):
          continue
        try:
          _, created = upsert_whitelist_user(connection, payload)
          if created:
            created_count += 1
          else:
            updated_count += 1
        except ValueError as error:
          skipped_rows.append({"row": row_number, "reason": str(error)})
      connection.commit()
  finally:
    workbook_path.unlink(missing_ok=True)

  return {
    "ok": True,
    "message": f"导入完成：新增 {created_count} 条，更新 {updated_count} 条，跳过 {len(skipped_rows)} 条。",
    "created": created_count,
    "updated": updated_count,
    "skipped": skipped_rows,
  }


@app.patch("/api/admin/users/{user_id}/status")
def update_user_status(user_id: int, payload: UserStatusPayload) -> dict[str, Any]:
  with get_connection() as connection:
    row = connection.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row:
      raise HTTPException(status_code=404, detail="用户不存在")
    connection.execute(
      "UPDATE users SET status = ?, updated_at = ? WHERE id = ?",
      (payload.status, now_iso(), user_id),
    )
    connection.commit()
    updated = connection.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    return user_to_admin(updated)


@app.get("/api/admin/layers")
def list_admin_layers(enabled_only: bool = Query(default=False)) -> list[dict[str, Any]]:
  query = "SELECT * FROM layers"
  params: tuple[Any, ...] = ()
  if enabled_only:
    query += " WHERE enabled = ?"
    params = (1,)
  query += " ORDER BY sort_order ASC, id ASC"
  with get_connection() as connection:
    rows = connection.execute(query, params).fetchall()
    return [layer_to_public(row) for row in rows]


@app.patch("/api/admin/layers/{layer_id}")
def update_layer(layer_id: str, payload: LayerPatchPayload) -> dict[str, Any]:
  updates: list[str] = []
  values: list[Any] = []
  field_map = {
    "name": ("name", payload.name),
    "categoryName": ("category_name", payload.categoryName),
    "enabled": ("enabled", None if payload.enabled is None else int(payload.enabled)),
    "sortOrder": ("sort_order", payload.sortOrder),
    "note": ("note", payload.note),
  }
  for _, (column, value) in field_map.items():
    if value is not None:
      updates.append(f"{column} = ?")
      values.append(value)

  if not updates:
    raise HTTPException(status_code=400, detail="没有可更新的字段")

  updates.append("updated_at = ?")
  values.append(now_iso())
  values.append(layer_id)

  with get_connection() as connection:
    row = connection.execute("SELECT * FROM layers WHERE id = ?", (layer_id,)).fetchone()
    if not row:
      raise HTTPException(status_code=404, detail="图层不存在")
    connection.execute(f"UPDATE layers SET {', '.join(updates)} WHERE id = ?", values)
    connection.commit()
    updated = connection.execute("SELECT * FROM layers WHERE id = ?", (layer_id,)).fetchone()
    return layer_to_public(updated)
